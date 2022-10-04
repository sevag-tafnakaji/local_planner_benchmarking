#!/usr/bin/env python3

from math import sqrt
from copy import deepcopy
import rospy
import os
from nav_msgs.msg import Path
from mobile_manipulator.msg import travelInfo, JointStates
from std_msgs.msg import Float32
import pandas as pd
from actionlib import GoalStatusArray
from geometry_msgs.msg import PoseStamped
from openpyxl import load_workbook
from gazebo_msgs.srv import GetModelState, GetModelStateRequest

# global variables
goal_pose = PoseStamped()
travel_info = travelInfo()
joint_states = JointStates()
global_times = []
local_times = []
real_path = []
global_planner_plan = []
status = 0
distance_travelled = Float32
path_area = Float32



def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def equalize_arr_length(arr1, arr2):
    length1 = len(arr1)
    length2 = len(arr2)
    if length1 < length2:
        for i in range(length1, length2):
            arr1.append(None)
    else:
        for i in range(length2, length1):
            arr2.append(None)

def get_goal_pose(goalPose):
    global goal_pose
    goal_pose = deepcopy(goalPose)


# Callback function
def count(msg):
    global global_planner_plan, global_times
    for element in msg.poses:
        global_planner_plan.append([element.pose.position.x, element.pose.position.y])
        global_times.append(element.header.stamp.to_sec())

def update_status(curr_status):
    global status
    for goal_status in curr_status.status_list:
        status = deepcopy(goal_status.status)

def calc_distance(pose1, pose2):
    x1 = pose1.pose.position.x
    x2 = pose2.pose.position.x
    y1 = pose1.pose.position.y
    y2 = pose2.pose.position.y
    dist = sqrt((x1-x2)**2+(y1-y2)**2)
    return dist

def update_real_path(Path):
    global real_path, local_times
    copied = deepcopy(Path)
    real_path = [[element.pose.position.x, element.pose.position.y] for element in copied.poses]
    local_times = [pose.header.stamp.to_sec() for pose in copied.poses]

def update_travel_info(info):
    global travel_info
    travel_info = deepcopy(info)
    
def update_joint_states(data):
    global joint_states
    joint_states = deepcopy(data)

def update_distance_travelled(dist_travelled):
    global distance_travelled
    distance_travelled = deepcopy(dist_travelled.data)

def update_area_between_paths(path_between_area):
    global path_area
    path_area = deepcopy(path_between_area.data)

if __name__ == "__main__":
    rospy.init_node("Navigation_stack_tester")
    # to keep track of navigation status
    rospy.Subscriber("/move_base/status", GoalStatusArray, update_status)

    # to get the info for the wobbliness of the path at the gripper
    rospy.Subscriber("gripper_info_publisher", travelInfo, update_travel_info, queue_size=5)

    # the real path taken by the robot
    rospy.Subscriber("/path", Path, update_real_path)

    # value for the total distance travelled
    rospy.Subscriber("/distance_metric", Float32, update_distance_travelled)

    # value for the area between the planned path and the actual path
    rospy.Subscriber("/area_between_paths", Float32, update_area_between_paths)

    # Depending on wich local and global planner you are using, rospy will subscribe to the good one and let the other.
    rospy.Subscriber("/move_base_node/GlobalPlanner/plan", Path, count)
    rospy.Subscriber("/move_base_node/SBPLLatticePlanner/plan", Path, count)

    # get joint states
    rospy.Subscriber("/arm_joints_publisher", JointStates, update_joint_states, queue_size=10)

    # getting the name of the launched world, planning alg, and trial number 
    # as these will affect where the data is stored
    world_name = rospy.get_param("/world_name_py")
    local_planner_alg = rospy.get_param("/local_planning_alg")
    local_planner_alg = local_planner_alg.upper()
    trial_number = rospy.get_param("trial_number")
    is_dynamic = rospy.get_param("is_dynamic")
    move_arm = rospy.get_param("is_moving")

    # writing the file paths for each metrics
    if(is_dynamic):
        file_path_ee = os.path.abspath("benchmarking/data/"+local_planner_alg+"/"+world_name+"/benchmarking_ee_path_"+world_name+"_dynamic.xlsx")
        file_paths_paths = os.path.abspath("benchmarking/data/"+local_planner_alg+"/"+world_name+"/benchmarking_paths_"+world_name+"_dynamic.xlsx")
        file_path_others = os.path.abspath("benchmarking/data/"+local_planner_alg+"/"+world_name+"/benchmarking_other_"+world_name+"_dynamic.xlsx")
    else:
        file_path_ee = os.path.abspath("benchmarking/data/"+local_planner_alg+"/"+world_name+"/benchmarking_ee_path_"+world_name+".xlsx")
        file_paths_paths = os.path.abspath("benchmarking/data/"+local_planner_alg+"/"+world_name+"/benchmarking_paths_"+world_name+".xlsx")
        file_path_others = os.path.abspath("benchmarking/data/"+local_planner_alg+"/"+world_name+"/benchmarking_other_"+world_name+".xlsx")
    if move_arm:
        file_path_arm_joints = os.path.abspath("benchmarking/data/"+local_planner_alg+"/"+world_name+"/benchmarking_moved_joints_"+world_name+".xlsx")
        file_path_ee = os.path.abspath("benchmarking/data/"+local_planner_alg+"/"+world_name+"/benchmarking_moved_ee_path_"+world_name+".xlsx")
        file_paths_paths = os.path.abspath("benchmarking/data/"+local_planner_alg+"/"+world_name+"/benchmarking_moved_paths_"+world_name+".xlsx")
        file_path_others = os.path.abspath("benchmarking/data/"+local_planner_alg+"/"+world_name+"/benchmarking_moved_other_"+world_name+".xlsx")
    
    # used for getting the final pose of the robot
    model = GetModelStateRequest()
    get_model_srv = rospy.ServiceProxy('/gazebo/get_model_state', GetModelState)
    model.model_name='Mobile_Manipulator'

    # setup for benchmarking is done! now the actual process can begin
    rospy.loginfo("Beginning test")

    # waiting until a goal is published
    goal_pose = rospy.wait_for_message("/move_base_simple/goal", PoseStamped)

    # start timer once the goal is recieved
    starting_time = rospy.get_time()

    # do nothing here while the robot is driving
    while(status != 3): 
        if rospy.is_shutdown():
            raise rospy.exceptions.ROSInterruptException("Manual shutdown given")
        continue
            
    
    #if the navigation has successfully reached the end
    if status == 3:
        rospy.logwarn("Target reached! Collecting Data")
        end_time = rospy.get_time() 
        execution_time = end_time - starting_time
        # waiting 5 seconds until all the data is published to be collected
        rospy.sleep(5)   
        final_pose = get_model_srv(model)
        dist_offset = calc_distance(goal_pose, final_pose)

        # arrays must be equal length to be exported to excel
        # filling up empty spots with None's
        equalize_arr_length(global_planner_plan, real_path)
        equalize_arr_length(global_times, global_planner_plan)
        equalize_arr_length(local_times, real_path)

        # if, for some reason, the equalize_arr_length function doesn't work 
        # as inteneded, it will be caught by the ValueError
        try:
            # collecting the data into pandas dataframes
            paths_data = pd.DataFrame({"global time":global_times, "global path": global_planner_plan,"local_times":local_times, "real path": real_path})
            ee_path_data = pd.DataFrame({"time": travel_info.times, "pose": travel_info.poses, "velocities": travel_info.vels})
            usual_metrics_data = pd.DataFrame({"Time taken": execution_time, "distance offset": dist_offset, "distance traveled": distance_travelled, "area between paths": path_area}, index=[0])
            joint_states_data = pd.DataFrame({"time": joint_states.times, "joint states": joint_states.states})
            
            rospy.logwarn("Data collected!")
            rospy.logwarn("Saving data into excel sheets")
            
            sheet_name = "Trial No."+ str(trial_number)
            files = {"ee_path": (file_path_ee, ee_path_data), "path": (file_paths_paths, paths_data), "other": (file_path_others, usual_metrics_data)}
            if move_arm:
                files = {"ee_path": (file_path_ee, ee_path_data), "path": (file_paths_paths, paths_data), "joints":(file_path_arm_joints, joint_states_data), "other": (file_path_others, usual_metrics_data)}
                print(files)
            # exporting the dataframes into their respective excel sheets
            for key, value in files.items():
                # Each algorithm will also have its own folder so make sure that the path updates to reflect that
                file_path = value[0]
                data = value[1]
                append_df_to_excel(filename=file_path,df=data, sheet_name=sheet_name, truncate_sheet=True)
            
            # end of this trial of benchmarking
            rospy.logwarn("Data saved, shutting down")
            rospy.set_param("/benchmarking_done", "true")
            rospy.signal_shutdown("Program did what it needed to do, shutting down.")
        
        except ValueError as e:
            # another script will catch this error and repeat the same trial
            print("program failed due to "+repr(e))
            rospy.set_param("/benchmarking_done", "failed")

    # If for some reason the navigation does not reach the goal
    else:
        # another script will catch this error and repeat the same trial
        rospy.logerr("Program failed to collect data. repeating trial")
        rospy.set_param("/benchmarking_done", "failed")
        rospy.signal_shutdown("Program failed, shutting down.")
        
    rospy.spin()